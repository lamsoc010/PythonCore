import tkinter as tk
from tkinter import filedialog
import openpyxl
from PIL import Image, ImageDraw, ImageFont

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        # Tạo label để yêu cầu người dùng chọn file Excel
        self.excel_label = tk.Label(self, text="Chọn file Excel")
        self.excel_label.pack(side="top")

        # Tạo button để cho người dùng chọn file Excel
        self.excel_button = tk.Button(self, text="Chọn file", command=self.select_excel)
        self.excel_button.pack(side="top")

        # Tạo label để yêu cầu người dùng chọn mẫu thẻ sinh viên
        self.template_label = tk.Label(self, text="Chọn mẫu thẻ sinh viên")
        self.template_label.pack(side="top")

        # Tạo listbox để cho người dùng chọn mẫu thẻ sinh viên
        self.template_listbox = tk.Listbox(self, height=3)
        self.template_listbox.pack(side="top")
        self.template_listbox.insert(1, "Mẫu 1")
        self.template_listbox.insert(2, "Mẫu 2")
        self.template_listbox.insert(3, "Mẫu 3")

        # Tạo button để bắt đầu tạo thẻ sinh viên
        self.create_button = tk.Button(self, text="Tạo thẻ sinh viên", command=self.create_cards)
        self.create_button.pack(side="bottom")

    def select_excel(self):
        # Mở hộp thoại để cho người dùng chọn file Excel
        self.excel_file = filedialog.askopenfilename()

    def create_cards(self):
        # Lấy thông tin file Excel và mẫu thẻ được chọn bởi người dùng
        excel_file = self.excel_file
        template = self.template_listbox.get(self.template_listbox.curselection())

        # Gọi hàm để tạo thẻ sinh viên
        create_student_cards(excel_file, template)

# Hàm để tạo thẻ sinh viên
def create_student_cards(excel_file, template):
    # Code để tạo thẻ sinh viên từ file Excel và mẫu thẻ được chọn
    # Mở file Excel và chọn sheet để đọc dữ liệu
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Lấy thông tin về số lượng sinh viên và các trường thông tin cần hiển thị trên thẻ sinh viên
    num_students = sheet.max_row - 1
    headers = [cell.value for cell in sheet[1]]

    # Load mẫu thẻ sinh viên
    template_file = f"{template}.png"
    card = Image.open(template_file)

    # Đặt font và kích thước cho các trường thông tin trên thẻ sinh viên
    font = ImageFont.truetype("arial.ttf", 16)
    text_color = (0, 0, 0)

    # Vẽ các trường thông tin lên thẻ sinh viên cho từng sinh viên
    for i in range(2, sheet.max_row + 1):
        # Lấy thông tin của một sinh viên
        student_info = [cell.value for cell in sheet[i]]

        # Tạo một hình ảnh mới từ mẫu thẻ sinh viên
        new_card = card.copy()

        # Vẽ các trường thông tin lên hình ảnh thẻ sinh viên
        draw = ImageDraw.Draw(new_card)
        for j, header in enumerate(headers):
            x = 50
            y = 100 + j * 30
            draw.text((x, y), f"{header}: {student_info[j]}", font=font, fill=text_color)

        # Lưu hình ảnh thẻ sinh viên mới
        new_card.save(f"card_{i-1}.png")

def show_student_card(student_id):
    # Tạo đường dẫn đến hình ảnh thẻ sinh viên
    image_path = f"card_{student_id}.png"
    
    # Kiểm tra xem hình ảnh có tồn tại không
    if not os.path.exists(image_path):
        print(f"Hình ảnh thẻ sinh viên {student_id} không tồn tại.")
        return
    
    # Đọc hình ảnh thẻ sinh viên và hiển thị nó
    image = Image.open(image_path)
    image.show(title=f"Thẻ sinh viên {student_id}")
    
root = tk.Tk()
app = Application(master=root)
app.mainloop()